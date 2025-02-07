import time

import flet as ft
import openpyxl


def main(page: ft.Page):
    # add records form
    def add_record_form(e):
        dialog_box.title = ft.Text(
            value="Add Records",
            weight=ft.FontWeight.BOLD
        )
        dialog_box.content = ft.Column(
            controls=[
                # Sr No label and textfield
                ft.Text(" Sr No"),
                ft.CupertinoTextField(
                    value=str(sh.max_row),
                    color=ft.Colors.GREY_400,
                    read_only=True,
                ),

                # space between Sr No and Goods Description
                ft.Container(height=5),

                # Goods Description label and textfield
                ft.Text(" Goods Description"),
                ft.CupertinoTextField(
                    placeholder_text="Goods Description",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    )
                ),

                # space between Goods Description and Part Number
                ft.Container(height=5),

                # Part Number label and textfield
                ft.Text(" Part Number"),
                ft.CupertinoTextField(
                    placeholder_text="Part Number",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    )
                ),

                # space between Part Number and Purchase Rate
                ft.Container(height=5),

                # Purchase Rate label and textfield
                ft.Text(" Purchase Rate"),
                ft.CupertinoTextField(
                    placeholder_text="Purchase Rate",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),

                # space between Purchase Rate and Stock
                ft.Container(height=5),

                # Stock label and textfield
                ft.Text(" Stock"),
                ft.CupertinoTextField(
                    placeholder_text="Stock",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),

                # space between Stock and HSN Code
                ft.Container(height=5),

                # HSN Code label and textfield
                ft.Text(" HSN Code"),
                ft.CupertinoTextField(
                    placeholder_text="HSN Code",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),

                # space between HSN Code and GST
                ft.Container(height=5),

                # GST label and textfield
                ft.Text(" GST"),
                ft.CupertinoTextField(
                    placeholder_text="GST",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),
            ]
        )

        # add button to add record to products_db.xlsx file
        # first clear if add buttons are already added
        dialog_box.actions.clear()
        dialog_box.actions.append(
            ft.TextButton(
                text="Add",
                on_click=add_record_to_database
            )
        )
        dialog_box.actions.append(
            ft.TextButton(
                text="Close",
                on_click=lambda e: page.close(dialog_box)
            )
        )

        page.open(dialog_box)

    # add record button
    def add_record_to_database(e):
        # add record to particular row (means after last row)
        select_row = sh.max_row + 1
        # Sr No
        sh.cell(row=select_row, column=1).value = int(dialog_box.content.controls[1].value)

        # Goods Description
        sh.cell(row=select_row, column=2).value = str(dialog_box.content.controls[4].value)

        # Part Number
        if str(dialog_box.content.controls[7].value).isdigit():
            sh.cell(row=select_row, column=3).value = int(dialog_box.content.controls[7].value)
        else:
            sh.cell(row=select_row, column=3).value = str(dialog_box.content.controls[7].value)

        # Purchase Rate
        sh.cell(row=select_row, column=4).value = int(dialog_box.content.controls[10].value)

        # Stock
        sh.cell(row=select_row, column=5).value = int(dialog_box.content.controls[13].value)

        # HSN Code
        sh.cell(row=select_row, column=6).value = int(dialog_box.content.controls[16].value)

        # GST
        sh.cell(row=select_row, column=7).value = int(dialog_box.content.controls[19].value)

        # save products_db
        wb.save("products_db.xlsx")

        # insert done icon at o index in dialog_box actions
        dialog_box.actions.insert(
            0,
            ft.Icon(
                name=ft.Icons.DONE,
                color=ft.Colors.GREEN
            )
        )

        # to update the new data record on the page
        add_data_row()

        page.update()

        # wait for one second before closing dialog box
        time.sleep(1)

        # close dialog_box
        page.close(dialog_box)
        # clear all button from dialog_box
        dialog_box.actions.clear()

    # control visibility of purchase rate column
    def control_visibility_of_purchase_rate(e):
        if e.control.icon == ft.Icons.VISIBILITY_OFF:
            e.control.icon = ft.Icons.VISIBILITY
            e.control.icon_color = ft.Colors.GREEN
            add_data_row()
        else:
            e.control.icon = ft.Icons.VISIBILITY_OFF
            e.control.icon_color = ft.Colors.RED
            add_data_row()

        page.update()

    # view record on single click of cell
    def view_record(e):
        dialog_box.title = ft.Text(
            value=f"Record No {e.control.data[0]}"
        )
        dialog_box.content = ft.Column(
            controls=[
                # Goods Description heading and value
                ft.Text(
                    value="Goods Description",
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[1],
                    color=ft.Colors.GREY
                ),

                # Part Number heading and value
                ft.Text(
                    value="Part Number",
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[2],
                    color=ft.Colors.GREY
                ),

                # Purchase Rate heading and value
                ft.Text(
                    value="Purchase Rate",
                    weight=ft.FontWeight.W_600,
                    visible=False
                ),
                ft.Text(
                    value=e.control.data[3],
                    color=ft.Colors.GREY,
                    visible=False
                ),

                # Sale Rate heading and value
                ft.Text(
                    value="Sale Rate",
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[4],
                    color=ft.Colors.GREY
                ),

                # Stock heading and value
                ft.Text(
                    value="Stock",
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[5],
                    color=ft.Colors.GREY
                ),

                # HSN Code heading and value
                ft.Text(
                    value="HSN Code",
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[6],
                    color=ft.Colors.GREY
                ),

                # GST heading and value
                ft.Text(
                    value="GST",
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[7],
                    color=ft.Colors.GREY
                ),

                # Total heading and value
                ft.Text(
                    value="Total",
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[8],
                    color=ft.Colors.GREY
                ),
            ]
        )

        # if purchase rate visibility is hidden than here also it is hidden else visible
        if visibility_purchase_btn.icon == ft.Icons.VISIBILITY_OFF:
            #  purchase rate heading visibility is hidden in view_record
            dialog_box.content.controls[4].visible = False
            #  purchase rate content visibility is hidden in view_record
            dialog_box.content.controls[5].visible = False
        else:
            #  purchase rate heading visibility is visible in view_record
            dialog_box.content.controls[4].visible = True
            #  purchase rate content visibility is visible in view_record
            dialog_box.content.controls[5].visible = True

        dialog_box.actions.clear()
        dialog_box.actions.append(
            ft.TextButton(
                text="Close",
                on_click=lambda e: page.close(dialog_box)
            )
        )
        page.open(dialog_box)

    # on long press to edit the record
    def edit_record(e):
        # Sr No, Sale Rate and Total columns are not editable
        if e.control.data[9] == "Sr No" or e.control.data[9] == "Sale Rate" or e.control.data[9] == "Total":
            dialog_box.title = ft.Text(
                value=e.control.data[9]
            )

            dialog_box.content = ft.Text(
                value=f"You can't update {e.control.data[9]}"
            )

            dialog_box.actions.clear()
            dialog_box.actions.append(
                ft.TextButton(
                    text="Close",
                    on_click=lambda e: page.close(dialog_box)
                )
            )

            page.open(dialog_box)
        else:
            dialog_box.title = ft.Text(
                value=f"Update Record No {e.control.data[0]}"
            )

            # this is the value for purchase column to avoid rupee symbol in textfield of editing
            val = ""
            if e.control.data[9] == "Purchase Rate":
                print("yes")
                val = str(e.control.content.value)[0:str(e.control.content.value).find("\u20B9") - 1]
            else:
                val = e.control.content.value
            dialog_box.content = ft.CupertinoTextField(
                value=val
            )

            dialog_box.actions.clear()
            dialog_box.actions.append(
                ft.TextButton(
                    text="Update",
                    data=e,
                    on_click=update_record
                )
            )
            dialog_box.actions.append(
                ft.TextButton(
                    text="Close",
                    on_click=lambda e: page.close(dialog_box)
                )
            )

            page.open(dialog_box)

    def update_record(e):
        # update the cell
        if str(dialog_box.content.value).isdigit():
            # if string contains only digits
            sh.cell(row=int(e.control.data.control.data[0]) + 1,
                    column=int(e.control.data.control.data[10])).value = int(dialog_box.content.value)
        else:
            # if string contains alphanumeric
            sh.cell(row=int(e.control.data.control.data[0]) + 1,
                    column=int(e.control.data.control.data[10])).value = str(dialog_box.content.value)

        # save Excel file
        wb.save("products_db.xlsx")

        # add done icon to show that record is updated
        dialog_box.actions.insert(
            0,
            ft.Icon(
                name=ft.Icons.DONE,
                color=ft.Colors.GREEN
            )
        )

        # to show updates call add_data_row function
        add_data_row()

        page.update()

        # wait for 1 second
        time.sleep(1)

        page.close(dialog_box)
        dialog_box.actions.clear()

    # search validations
    def search_field_validation(e):
        # if search box is having some data than close btn appears in the search box
        if e.control.value != "":
            e.control.suffix.content.name = ft.Icons.CLOSE_ROUNDED
        else:
            e.control.suffix.content.name = None

        add_data_row()

        page.update()

    # clear search field
    def clear_search_field(e):
        # when clear btn pressed search box clear the search box content and hide the close btn
        e.control.content.name = None
        search_box.value = ""

        add_data_row()

        page.update()

    # page settings
    page.padding = 0
    page.spacing = 0
    page.bgcolor = ft.Colors.GREY_100
    page.theme_mode = ft.ThemeMode.LIGHT

    # alert dialog box
    dialog_box = ft.AlertDialog(
        # dialog box background color
        surface_tint_color=ft.Colors.WHITE,
        scrollable=True,
        modal=True,
    )

    # visibility btn to control visibility of purchase rate column
    visibility_purchase_btn = ft.IconButton(
        icon=ft.Icons.VISIBILITY_OFF,
        icon_color=ft.Colors.RED,
        on_click=control_visibility_of_purchase_rate
    )

    # page appbar
    page.appbar = ft.AppBar(
        adaptive=True,
        # row for home icon and products title
        title=ft.Row(
            spacing=10, run_spacing=0,
            controls=[
                # home icon
                ft.Icon(
                    name=ft.Icons.HOME,
                    color=ft.Colors.WHITE,
                    size=30
                ),

                # products title
                ft.Text(
                    value="PRODUCTS",
                    color=ft.Colors.WHITE,
                    weight=ft.FontWeight.BOLD
                )
            ]
        ),

        # app background color
        bgcolor="#36618E",

        # action button on the end of appbar
        actions=[
            # menu button
            ft.PopupMenuButton(
                icon=ft.Icons.MENU_OPEN,
                icon_color=ft.Colors.WHITE,
                items=[
                    # menu heading
                    ft.PopupMenuItem(
                        content=ft.Text(
                            value="Menu",
                            weight=ft.FontWeight.BOLD
                        )

                    ),

                    # divider
                    ft.PopupMenuItem(),

                    # go to quotation page btn
                    ft.PopupMenuItem(
                        content=ft.Row(
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            controls=[
                                # quotation title
                                ft.Text(
                                    value="View Quotation"
                                ),

                                # icon button to show quotation
                                ft.IconButton(
                                    icon=ft.Icons.REQUEST_QUOTE,
                                    icon_color=ft.Colors.GREEN,
                                    on_click=None
                                )
                            ]
                        )
                    ),

                    # divider
                    ft.PopupMenuItem(),

                    # add records to products table button
                    ft.PopupMenuItem(
                        content=ft.Row(
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            controls=[
                                # add records title
                                ft.Text(
                                    value="Add New Record"
                                ),

                                # icon button to show add record form
                                ft.IconButton(
                                    icon=ft.Icons.ADD_BOX,
                                    icon_color=ft.Colors.GREEN,
                                    on_click=add_record_form
                                )
                            ]
                        )
                    ),

                    # divider
                    ft.PopupMenuItem(),

                    # control visibility of purchase rate column
                    ft.PopupMenuItem(
                        content=ft.Row(
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            controls=[
                                # visible purchase rate title
                                ft.Text(
                                    value="Visible Purchase Rate"
                                ),

                                # icon button to control visibility of purchase rate column
                                visibility_purchase_btn,
                            ]
                        )

                    ),

                    # divider
                    ft.PopupMenuItem(),

                    # Information
                    ft.PopupMenuItem(
                        content=ft.Column(
                            spacing=20, run_spacing=0,
                            controls=[
                                # single click instruction
                                ft.Text(
                                    value="Tap to view Record",
                                    color=ft.Colors.GREY,
                                ),

                                # double click instruction
                                ft.Text(
                                    value="Long press to edit Record",
                                    color=ft.Colors.GREY,
                                )
                            ]
                        )
                    ),
                ]
            )
        ]
    )

    # search box
    search_box = ft.CupertinoTextField(
        placeholder_text="Search",
        border_radius=0,
        padding=ft.padding.only(
            left=20, top=15, bottom=15
        ),
        border=ft.border.only(
            bottom=ft.BorderSide(width=1, color=ft.Colors.GREY_300)
        ),
        suffix=ft.Container(
            ft.Icon(
                name=None,
                color=ft.Colors.RED
            ),
            bgcolor=ft.Colors.WHITE,
            padding=ft.padding.only(right=20),
            on_click=clear_search_field,
        ),
        on_change=search_field_validation
    )

    # data row
    data_row = ft.Column(
        spacing=1, run_spacing=0,
        scroll=ft.ScrollMode.HIDDEN,
        height=690
    )

    # products table container
    products_table = ft.Row(
        spacing=0, run_spacing=0,
        scroll=ft.ScrollMode.HIDDEN,
        controls=[
            ft.Column(
                spacing=1, run_spacing=0,
                controls=[
                    ft.Row(
                        spacing=1,
                        controls=[
                            # quote column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=100,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Quote",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # Sr No column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=100,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Sr No",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # goods description column
                            ft.Container(
                                padding=ft.padding.only(left=10, top=10, bottom=10),
                                width=300,
                                alignment=ft.alignment.center_left,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Goods Description",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # part number column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=200,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Part Number",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # purchase rate column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=150,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Purchase rate",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # sale rate column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=150,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Sale rate",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # stock column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=100,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Stock",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # hsn code column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=200,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="HSN Code",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # gst column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=100,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="GST",
                                    weight=ft.FontWeight.BOLD
                                )
                            ),

                            # total column
                            ft.Container(
                                padding=ft.padding.only(top=10, bottom=10),
                                width=150,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREY_300,
                                content=ft.Text(
                                    value="Total",
                                    weight=ft.FontWeight.BOLD
                                )
                            )
                        ]
                    ),

                    data_row
                ]
            )
        ]
    )

    # load workbook
    wb = openpyxl.load_workbook("products_db.xlsx")
    # select sheet
    sh = wb["Sheet1"]

    def add_data_row():
        # clear data_row before append
        data_row.controls.clear()

        for i in range(2, sh.max_row + 1):
            # sale rate calculation from purchase rate
            sale_rate = round(int(sh.cell(row=i, column=4).value) * 0.1 + int(sh.cell(row=i, column=4).value))

            # append data to data_row
            data_row.controls.append(
                ft.Row(
                    expand=True,
                    spacing=1, run_spacing=0,
                    controls=[
                        # quote column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Container(
                                width=25,
                                height=25,
                                border_radius=5,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREEN,
                                content=ft.Icon(
                                    name=ft.Icons.ADD,
                                    color=ft.Colors.WHITE
                                )
                            )
                        ),

                        # Sr No column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(sh.cell(row=i, column=1).value)
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Sr No",  # column name
                                "1",  # column number in Excel file
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # goods description column
                        ft.Container(
                            width=300,
                            height=60,
                            padding=ft.padding.only(left=10, right=10),
                            alignment=ft.alignment.center_left,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(sh.cell(row=i, column=2).value)
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Goods Description",  # column name
                                "2"  # column number in Excel file
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # part number column
                        ft.Container(
                            width=200,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(sh.cell(row=i, column=3).value)
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Part Number",  # column name
                                "3"  # column number in Excel file
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # purchase rate column
                        ft.Container(
                            width=150,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{sh.cell(row=i, column=4).value} \u20B9"
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Purchase Rate",  # column name
                                "4"  # column number in Excel file
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # sale rate column
                        ft.Container(
                            width=150,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{sale_rate} \u20B9"
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Sale Rate",  # column name
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # stock column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(sh.cell(row=i, column=5).value)
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Stock",  # column name
                                "5"  # column number in Excel file
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # hsn code column
                        ft.Container(
                            width=200,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(sh.cell(row=i, column=6).value)
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "HSN Code",  # column name
                                "6"  # column number in Excel file
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # gst column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{sh.cell(row=i, column=7).value}%"
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "GST",  # column name
                                "7",  # column number in Excel file
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        ),

                        # total column
                        ft.Container(
                            width=150,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{round(sale_rate * 0.18 + sale_rate)} \u20B9"
                            ),
                            data=[
                                str(sh.cell(row=i, column=1).value),  # Sr No value
                                str(sh.cell(row=i, column=2).value),  # Goods Description value
                                str(sh.cell(row=i, column=3).value),  # Part number value
                                f"{sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(sh.cell(row=i, column=5).value),  # Stock value
                                str(sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Total",  # column name
                            ],
                            on_click=view_record,
                            on_long_press=edit_record
                        )
                    ]
                )
            )

            # initial purchase rate is hidden
            # if we click on eye btn in appbar than it will be visible
            if visibility_purchase_btn.icon == ft.Icons.VISIBILITY_OFF:
                # heading of Purchase Rate visibility False
                products_table.controls[0].controls[0].controls[4].visible = False
                # content of Purchase Rate visibility False
                data_row.controls[i - 2].controls[4].visible = False
            else:
                # heading of Purchase Rate visibility True
                products_table.controls[0].controls[0].controls[4].visible = True
                # content of Purchase Rate visibility True
                data_row.controls[i - 2].controls[4].visible = True

            # if stock is zero than show it in red color
            if sh.cell(row=i, column=5).value == 0:
                data_row.controls[i-2].controls[6].content.color = ft.Colors.RED

            # if no part number available show empty cell
            if sh.cell(row=i, column=3).value is None:
                data_row.controls[i - 2].controls[3].content.value = ""

            # show only matched value with search box
            if search_box.value.lower() in str(sh.cell(row=i, column=2).value).lower():
                data_row.controls[i - 2].visible = True
                # data_row.controls[i - 2].controls[1].visible = True
                # data_row.controls[i - 2].controls[2].visible = True
                # data_row.controls[i - 2].controls[3].visible = True
                # data_row.controls[i - 2].controls[4].visible = True
                # data_row.controls[i - 2].controls[5].visible = True
                # data_row.controls[i - 2].controls[6].visible = True
                # data_row.controls[i - 2].controls[7].visible = True
                # data_row.controls[i - 2].controls[8].visible = True
                # data_row.controls[i - 2].controls[9].visible = True
            else:
                data_row.controls[i - 2].visible = False
                # data_row.controls[i - 2].controls[1].visible = False
                # data_row.controls[i - 2].controls[2].visible = False
                # data_row.controls[i - 2].controls[3].visible = False
                # data_row.controls[i - 2].controls[4].visible = False
                # data_row.controls[i - 2].controls[5].visible = False
                # data_row.controls[i - 2].controls[6].visible = False
                # data_row.controls[i - 2].controls[7].visible = False
                # data_row.controls[i - 2].controls[8].visible = False
                # data_row.controls[i - 2].controls[9].visible = False

    add_data_row()

    page.add(
        search_box,
        products_table
    )


ft.app(main)
