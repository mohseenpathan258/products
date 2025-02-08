import flet as ft
import openpyxl
import time
import os


class Products(ft.Column):
    def __init__(self, page):
        super(Products, self).__init__()

        self.page = page

        # used to add quotation data
        if self.page.client_storage.contains_key("quotation"):
            self.quotation_list = self.page.client_storage.get("quotation")
        else:
            self.quotation_list = []

        # load workbook
        self.file_path = "products_db.xlsx"
        self.wb = openpyxl.load_workbook(filename=os.path.abspath(self.file_path))
        # select sheet
        self.sh = self.wb["Sheet1"]

        # alert dialog box
        self.dialog_box = ft.AlertDialog(
            # dialog box background color
            surface_tint_color=ft.Colors.WHITE,
            scrollable=True,
            modal=True,
        )

        # snack bar
        self.snack_bar = ft.SnackBar(
            content=ft.Text()
        )

        # visibility btn to control visibility of purchase rate column
        self.visibility_purchase_btn = ft.IconButton(
            icon=ft.Icons.VISIBILITY_OFF,
            icon_color=ft.Colors.RED,
            on_click=self.control_visibility_of_purchase_rate
        )

        # custom app_bar
        self.app_bar = ft.ResponsiveRow(
            spacing=0, run_spacing=0,
            controls=[
                # app_bar container
                ft.Container(
                    height=70,
                    bgcolor=ft.Colors.YELLOW,
                    alignment=ft.alignment.center_left,
                    # padding=ft.padding.only(left=10, bottom=10, right=10),
                    content=ft.Row(
                        controls=[
                            # app_bar home button
                            ft.IconButton(
                                icon=ft.Icons.HOME,
                                icon_color=ft.Colors.BLACK,
                                icon_size=30,
                                on_click=lambda e: self.page.go("/products")
                            ),

                            # app_bar products title
                            ft.Text(
                                value="PRODUCTS",
                                color=ft.Colors.BLACK,
                                size=25,
                                weight=ft.FontWeight.BOLD
                            ),

                            ft.Row(
                                expand=True,
                                alignment=ft.MainAxisAlignment.END,
                                controls=[
                                    # menu button
                                    ft.PopupMenuButton(
                                        icon=ft.Icons.MENU_OPEN,
                                        icon_size=30,
                                        icon_color=ft.Colors.BLACK,
                                        items=[
                                            # menu heading
                                            ft.PopupMenuItem(
                                                content=ft.Text(
                                                    value="Menu",
                                                    color=ft.Colors.BLACK,
                                                    weight=ft.FontWeight.BOLD
                                                )

                                            ),

                                            # divider
                                            ft.PopupMenuItem(),

                                            # go to quotation page btn
                                            ft.PopupMenuItem(
                                                content=ft.Row(
                                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                                                    controls=[
                                                        # quotation title
                                                        ft.Text(
                                                            value="View Quotation",
                                                            color=ft.Colors.BLACK
                                                        ),

                                                        # icon button to show quotation
                                                        ft.IconButton(
                                                            icon=ft.Icons.REQUEST_QUOTE,
                                                            icon_color=ft.Colors.GREEN,
                                                            on_click=lambda e: self.page.go("/quotation")
                                                        )
                                                    ]
                                                )
                                            ),

                                            # divider
                                            ft.PopupMenuItem(),

                                            # clear quotation
                                            ft.PopupMenuItem(
                                                content=ft.Row(
                                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                                                    controls=[
                                                        # clear quotation title
                                                        ft.Text(
                                                            value="Clear Quotation",
                                                            color=ft.Colors.BLACK
                                                        ),

                                                        # icon button to clear quotation
                                                        ft.IconButton(
                                                            icon=ft.Icons.REQUEST_QUOTE,
                                                            icon_color=ft.Colors.RED,
                                                            on_click=self.clear_quotation
                                                        )
                                                    ]
                                                )
                                            ),

                                            # divider
                                            ft.PopupMenuItem(),

                                            # add new record title and button
                                            ft.PopupMenuItem(
                                                content=ft.Row(
                                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                                                    controls=[
                                                        # add records title
                                                        ft.Text(
                                                            value="Add New Record",
                                                            color=ft.Colors.BLACK,
                                                        ),

                                                        # icon button to show add record form
                                                        ft.IconButton(
                                                            icon=ft.Icons.ADD_BOX,
                                                            icon_color=ft.Colors.GREEN,
                                                            on_click=self.add_record_form
                                                        )
                                                    ]
                                                )
                                            ),

                                            # divider
                                            ft.PopupMenuItem(),

                                            # visible purchase rate title and button
                                            # control the visibility of purchase rate column
                                            ft.PopupMenuItem(
                                                content=ft.Row(
                                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                                                    controls=[
                                                        # visible purchase rate title
                                                        ft.Text(
                                                            value="Visible Purchase Rate",
                                                            color=ft.Colors.BLACK,
                                                        ),

                                                        # icon button to control visibility of purchase rate column
                                                        self.visibility_purchase_btn,
                                                    ]
                                                )

                                            ),

                                            # divider
                                            ft.PopupMenuItem(),

                                            # information of "on tap" and "on long press"
                                            ft.PopupMenuItem(
                                                content=ft.Column(
                                                    spacing=20, run_spacing=0,
                                                    controls=[
                                                        # information title
                                                        ft.Text(
                                                            value="Information",
                                                            color=ft.Colors.BLACK,
                                                            weight=ft.FontWeight.W_600
                                                        ),

                                                        # single click instruction
                                                        ft.Text(
                                                            value="Tap to view Record",
                                                            color=ft.Colors.GREY,
                                                        ),

                                                        # double click instruction
                                                        ft.Text(
                                                            value="Long press to edit Record",
                                                            color=ft.Colors.GREY,
                                                        )
                                                    ]
                                                )
                                            ),

                                            # divider
                                            ft.PopupMenuItem(),

                                            # logout title and button
                                            ft.PopupMenuItem(
                                                content=ft.Row(
                                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                                                    controls=[
                                                        # add records title
                                                        ft.Text(
                                                            value="Logout",
                                                            color=ft.Colors.GREY_700,
                                                        ),

                                                        # icon button to show add record form
                                                        ft.IconButton(
                                                            icon=ft.Icons.LOGOUT,
                                                            icon_color=ft.Colors.RED,
                                                            on_click=lambda e: self.page.go("/login")
                                                        )
                                                    ]
                                                )
                                            ),
                                        ]
                                    )
                                ]
                            )
                        ]
                    )
                )
            ]
        )

        # search box
        self.search_box = ft.CupertinoTextField(
            placeholder_text="Search",
            placeholder_style=ft.TextStyle(color=ft.Colors.GREY),
            color=ft.Colors.BLACK,
            bgcolor=ft.Colors.WHITE,
            border_radius=0,
            padding=ft.padding.only(
                left=20, top=15, bottom=15
            ),
            border=ft.border.only(
                bottom=ft.BorderSide(width=1, color=ft.Colors.GREY_300)
            ),
            suffix=ft.Container(
                ft.Icon(
                    name=None,
                    color=ft.Colors.RED
                ),
                bgcolor=ft.Colors.WHITE,
                padding=ft.padding.only(right=20),
                on_click=self.clear_search_field,
            ),
            on_change=self.search_field_validation
        )

        # data row
        self.data_row = ft.Column(
            spacing=1, run_spacing=0,
            scroll=ft.ScrollMode.HIDDEN,
            height=690
        )

        # products table container
        self.products_table = ft.Row(
            spacing=0, run_spacing=0,
            scroll=ft.ScrollMode.HIDDEN,
            controls=[
                ft.Column(
                    spacing=1, run_spacing=0,
                    controls=[
                        ft.Row(
                            spacing=1,
                            controls=[
                                # quote column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=100,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Quote",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # Sr No column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=100,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Sr No",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # goods description column
                                ft.Container(
                                    padding=ft.padding.only(left=10, top=10, bottom=10),
                                    width=300,
                                    alignment=ft.alignment.center_left,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Goods Description",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # part number column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=200,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Part Number",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # purchase rate column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=150,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Purchase rate",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # sale rate column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=150,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Sale rate",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # stock column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=100,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Stock",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # hsn code column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=200,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="HSN Code",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # gst column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=100,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="GST",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                ),

                                # total column
                                ft.Container(
                                    padding=ft.padding.only(top=10, bottom=10),
                                    width=150,
                                    alignment=ft.alignment.center,
                                    bgcolor=ft.Colors.GREY_300,
                                    content=ft.Text(
                                        value="Total",
                                        color=ft.Colors.BLACK,
                                        weight=ft.FontWeight.BOLD
                                    )
                                )
                            ]
                        ),

                        self.data_row
                    ]
                )
            ]
        )

        self.spacing = 0
        self.controls = [
            self.app_bar,

            self.search_box,

            self.products_table,
        ]

        self.add_data_row()

    def add_data_row(self):
        # clear data_row before append
        self.data_row.controls.clear()

        for i in range(2, self.sh.max_row + 1):
            # sale rate calculation from purchase rate
            sale_rate = round(int(self.sh.cell(row=i, column=4).value) * 0.1 + int(self.sh.cell(row=i, column=4).value))

            # append data to data_row
            self.data_row.controls.append(
                ft.Row(
                    expand=True,
                    spacing=1, run_spacing=0,
                    controls=[
                        # quote column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Container(
                                width=25,
                                height=25,
                                border_radius=5,
                                alignment=ft.alignment.center,
                                bgcolor=ft.Colors.GREEN,
                                content=ft.Icon(
                                    name=ft.Icons.ADD,
                                    color=ft.Colors.WHITE,
                                ),
                                data=[
                                    str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                    str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                    str(self.sh.cell(row=i, column=3).value),  # Part number value
                                    f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                    f"{sale_rate} \u20B9",  # Sale rate value
                                    str(self.sh.cell(row=i, column=5).value),  # Stock value
                                    str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                    f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                    f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                    "Sr No",  # column name
                                    "1",  # column number in Excel file
                                ],
                                on_click=self.prepare_quotation
                            )
                        ),

                        # Sr No column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=1).value)
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Sr No",  # column name
                                "1",  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # goods description column
                        ft.Container(
                            width=300,
                            height=60,
                            padding=ft.padding.only(left=10, right=10),
                            alignment=ft.alignment.center_left,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=2).value)
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Goods Description",  # column name
                                "2"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # part number column
                        ft.Container(
                            width=200,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=3).value)
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Part Number",  # column name
                                "3"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # purchase rate column
                        ft.Container(
                            width=150,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{self.sh.cell(row=i, column=4).value} \u20B9"
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Purchase Rate",  # column name
                                "4"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # sale rate column
                        ft.Container(
                            width=150,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{sale_rate} \u20B9"
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Sale Rate",  # column name
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # stock column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=5).value)
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Stock",  # column name
                                "5"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # hsn code column
                        ft.Container(
                            width=200,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=6).value)
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "HSN Code",  # column name
                                "6"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # gst column
                        ft.Container(
                            width=100,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{self.sh.cell(row=i, column=7).value}%"
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "GST",  # column name
                                "7",  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # total column
                        ft.Container(
                            width=150,
                            height=60,
                            alignment=ft.alignment.center,
                            bgcolor=ft.Colors.WHITE,
                            content=ft.Text(
                                value=f"{round(sale_rate * 0.18 + sale_rate)} \u20B9"
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Total",  # column name
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        )
                    ]
                )
            )

            # initial purchase rate is hidden
            # if we click on eye btn in appbar than it will be visible
            if self.visibility_purchase_btn.icon == ft.Icons.VISIBILITY_OFF:
                # heading of Purchase Rate visibility False
                self.products_table.controls[0].controls[0].controls[4].visible = False
                # content of Purchase Rate visibility False
                self.data_row.controls[i - 2].controls[4].visible = False
            else:
                # heading of Purchase Rate visibility True
                self.products_table.controls[0].controls[0].controls[4].visible = True
                # content of Purchase Rate visibility True
                self.data_row.controls[i - 2].controls[4].visible = True

            # if stock is zero than show it in red color
            if self.sh.cell(row=i, column=5).value == 0:
                self.data_row.controls[i - 2].controls[6].content.color = ft.Colors.RED

            # if no part number available self.show empty cell
            if self.sh.cell(row=i, column=3).value is None:
                self.data_row.controls[i - 2].controls[3].content.value = ""

            # show only matched value with search box
            if self.search_box.value.lower() in str(self.sh.cell(row=i, column=2).value).lower():
                self.data_row.controls[i - 2].visible = True
            else:
                self.data_row.controls[i - 2].visible = False

            # to change the quotation btn to remove icon if quotation is not empty
            # to change the quotation btn to add icon if quotation is empty
            check_quotation_already_added = [self.data_row.controls[i - 2].controls[0].content.data[0], self.data_row.controls[i - 2].controls[0].content.data[1], self.data_row.controls[i - 2].controls[0].content.data[5], self.data_row.controls[i - 2].controls[0].content.data[4]]
            if check_quotation_already_added in self.quotation_list:
                self.data_row.controls[i - 2].controls[0].content.content.name = ft.Icons.REMOVE
                self.data_row.controls[i - 2].controls[0].content.bgcolor = ft.Colors.RED
                self.page.update()

    # search validations
    def search_field_validation(self, e):
        # if search box is having some data than close btn appears in the search box
        if e.control.value != "":
            e.control.suffix.content.name = ft.Icons.CLOSE_ROUNDED
        else:
            e.control.suffix.content.name = None

        self.add_data_row()

        self.page.update()

    # clear search field
    def clear_search_field(self, e):
        # when clear btn pressed search box clear the search box content and hide the close btn
        e.control.content.name = None
        self.search_box.value = ""

        self.add_data_row()

        self.page.update()

    # control visibility of purchase rate column
    def control_visibility_of_purchase_rate(self, e):
        if e.control.icon == ft.Icons.VISIBILITY_OFF:
            e.control.icon = ft.Icons.VISIBILITY
            e.control.icon_color = ft.Colors.GREEN
            self.add_data_row()
        else:
            e.control.icon = ft.Icons.VISIBILITY_OFF
            e.control.icon_color = ft.Colors.RED
            self.add_data_row()

        self.page.update()

    # add records form
    def add_record_form(self, e):
        self.dialog_box.title = ft.Text(
            value="Add Records",
            weight=ft.FontWeight.BOLD
        )
        self.dialog_box.content = ft.Column(
            controls=[
                # Sr No label and textfield
                ft.Text(" Sr No"),
                ft.CupertinoTextField(
                    value=str(self.sh.max_row),
                    color=ft.Colors.GREY_400,
                    read_only=True,
                ),

                # space between Sr No and Goods Description
                ft.Container(height=5),

                # Goods Description label and textfield
                ft.Text(" Goods Description"),
                ft.CupertinoTextField(
                    placeholder_text="Goods Description",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    )
                ),

                # space between Goods Description and Part Number
                ft.Container(height=5),

                # Part Number label and textfield
                ft.Text(" Part Number"),
                ft.CupertinoTextField(
                    placeholder_text="Part Number",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    )
                ),

                # space between Part Number and Purchase Rate
                ft.Container(height=5),

                # Purchase Rate label and textfield
                ft.Text(" Purchase Rate"),
                ft.CupertinoTextField(
                    placeholder_text="Purchase Rate",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),

                # space between Purchase Rate and Stock
                ft.Container(height=5),

                # Stock label and textfield
                ft.Text(" Stock"),
                ft.CupertinoTextField(
                    placeholder_text="Stock",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),

                # space between Stock and HSN Code
                ft.Container(height=5),

                # HSN Code label and textfield
                ft.Text(" HSN Code"),
                ft.CupertinoTextField(
                    placeholder_text="HSN Code",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),

                # space between HSN Code and GST
                ft.Container(height=5),

                # GST label and textfield
                ft.Text(" GST"),
                ft.CupertinoTextField(
                    placeholder_text="GST",
                    placeholder_style=ft.TextStyle(
                        color=ft.Colors.GREY_400,
                    ),
                    keyboard_type=ft.KeyboardType.NUMBER
                ),
            ]
        )

        # add button to add record to products_db.xlsx file
        # first clear if add buttons are already added
        self.dialog_box.actions.clear()
        self.dialog_box.actions.append(
            ft.TextButton(
                text="Add",
                on_click=self.add_record_to_database
            )
        )
        self.dialog_box.actions.append(
            ft.TextButton(
                text="Close",
                on_click=lambda event: self.page.close(self.dialog_box)
            )
        )

        self.page.open(self.dialog_box)

    # add record button
    def add_record_to_database(self, e):
        # add record to particular row (means after last row)
        select_row = self.sh.max_row + 1
        print(select_row)
        # Sr No
        self.sh.cell(row=select_row, column=1).value = int(self.dialog_box.content.controls[1].value)

        # Goods Description
        self.sh.cell(row=select_row, column=2).value = str(self.dialog_box.content.controls[4].value)

        # Part Number
        if str(self.dialog_box.content.controls[7].value).isdigit():
            self.sh.cell(row=select_row, column=3).value = int(self.dialog_box.content.controls[7].value)
        else:
            self.sh.cell(row=select_row, column=3).value = str(self.dialog_box.content.controls[7].value)

        # Purchase Rate
        self.sh.cell(row=select_row, column=4).value = int(self.dialog_box.content.controls[10].value)

        # Stock
        self.sh.cell(row=select_row, column=5).value = int(self.dialog_box.content.controls[13].value)

        # HSN Code
        self.sh.cell(row=select_row, column=6).value = int(self.dialog_box.content.controls[16].value)

        # GST
        self.sh.cell(row=select_row, column=7).value = int(self.dialog_box.content.controls[19].value)

        # save products_db
        self.wb.save(os.path.abspath(self.file_path))

        # insert done icon at o index in dialog_box actions
        self.dialog_box.actions.insert(
            0,
            ft.Icon(
                name=ft.Icons.DONE,
                color=ft.Colors.GREEN
            )
        )

        # to update the new data record on the page
        self.add_data_row()

        self.page.update()

        # wait for one second before closing dialog box
        time.sleep(1)

        # close dialog_box
        self.page.close(self.dialog_box)
        # clear all button from dialog_box
        self.dialog_box.actions.clear()

    # view record on single click of cell
    def view_record(self, e):
        self.dialog_box.title = ft.Text(
            value=f"Record No {e.control.data[0]}"
        )
        self.dialog_box.content = ft.Column(
            controls=[
                # Goods Description heading and value
                ft.Text(
                    value="Goods Description",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[1],
                    color=ft.Colors.GREY
                ),

                # Part Number heading and value
                ft.Text(
                    value="Part Number",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[2],
                    color=ft.Colors.GREY
                ),

                # Purchase Rate heading and value
                ft.Text(
                    value="Purchase Rate",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600,
                    visible=False
                ),
                ft.Text(
                    value=e.control.data[3],
                    color=ft.Colors.GREY,
                    visible=False
                ),

                # Sale Rate heading and value
                ft.Text(
                    value="Sale Rate",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[4],
                    color=ft.Colors.GREY
                ),

                # Stock heading and value
                ft.Text(
                    value="Stock",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[5],
                    color=ft.Colors.GREY
                ),

                # HSN Code heading and value
                ft.Text(
                    value="HSN Code",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[6],
                    color=ft.Colors.GREY
                ),

                # GST heading and value
                ft.Text(
                    value="GST",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[7],
                    color=ft.Colors.GREY
                ),

                # Total heading and value
                ft.Text(
                    value="Total",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[8],
                    color=ft.Colors.GREY
                ),
            ]
        )

        # if purchase rate visibility is hidden than here also it is hidden else visible
        if self.visibility_purchase_btn.icon == ft.Icons.VISIBILITY_OFF:
            #  purchase rate heading visibility is hidden in view_record
            self.dialog_box.content.controls[4].visible = False
            #  purchase rate content visibility is hidden in view_record
            self.dialog_box.content.controls[5].visible = False
        else:
            #  purchase rate heading visibility is visible in view_record
            self.dialog_box.content.controls[4].visible = True
            #  purchase rate content visibility is visible in view_record
            self.dialog_box.content.controls[5].visible = True

        self.dialog_box.actions.clear()
        self.dialog_box.actions.append(
            ft.TextButton(
                text="Close",
                on_click=lambda e: self.page.close(self.dialog_box)
            )
        )
        self.page.open(self.dialog_box)

        # on long press to edit the record

    def edit_record(self, e):
        # Sr No, Sale Rate and Total columns are not editable
        if e.control.data[9] == "Sr No" or e.control.data[9] == "Sale Rate" or e.control.data[9] == "Total":
            self.dialog_box.title = ft.Text(
                value=e.control.data[9]
            )

            self.dialog_box.content = ft.Text(
                value=f"You can't update {e.control.data[9]}"
            )

            self.dialog_box.actions.clear()
            self.dialog_box.actions.append(
                ft.TextButton(
                    text="Close",
                    on_click=lambda e: self.page.close(self.dialog_box)
                )
            )

            self.page.open(self.dialog_box)
        else:
            self.dialog_box.title = ft.Text(
                value=f"Update Record No {e.control.data[0]}"
            )

            # this is the value for purchase column to avoid rupee symbol in textfield of editing
            val = ""
            if e.control.data[9] == "Purchase Rate":
                val = str(e.control.content.value)[0:str(e.control.content.value).find("\u20B9") - 1]
            else:
                val = e.control.content.value
            self.dialog_box.content = ft.CupertinoTextField(
                value=val
            )

            self.dialog_box.actions.clear()
            self.dialog_box.actions.append(
                ft.TextButton(
                    text="Update",
                    data=e,
                    on_click=self.update_record
                )
            )
            self.dialog_box.actions.append(
                ft.TextButton(
                    text="Close",
                    on_click=lambda e: self.page.close(self.dialog_box)
                )
            )

            self.page.open(self.dialog_box)

    def update_record(self, e):
        # update the cell
        if str(self.dialog_box.content.value).isdigit():
            # if string contains only digits
            self.sh.cell(row=int(e.control.data.control.data[0]) + 1,
                         column=int(e.control.data.control.data[10])).value = int(self.dialog_box.content.value)
        else:
            # if string contains alphanumeric
            self.sh.cell(row=int(e.control.data.control.data[0]) + 1,
                         column=int(e.control.data.control.data[10])).value = str(self.dialog_box.content.value)

        # save Excel file
        self.wb.save(os.path.abspath(self.file_path))

        # add done icon to show that record is updated
        self.dialog_box.actions.insert(
            0,
            ft.Icon(
                name=ft.Icons.DONE,
                color=ft.Colors.GREEN
            )
        )

        # to show updates call add_data_row function
        self.add_data_row()

        self.page.update()

        # wait for 1 second
        time.sleep(1)

        self.page.close(self.dialog_box)
        self.dialog_box.actions.clear()

    # prepare quotation
    def prepare_quotation(self, e):
        if e.control.content.name == ft.Icons.ADD:
            quotation_list_row = [e.control.data[0], e.control.data[1], e.control.data[5], e.control.data[4]]
            self.quotation_list.append(quotation_list_row)
            self.page.client_storage.set("quotation", self.quotation_list)

            self.page.client_storage.set("add_remove_quotation_icon", ft.Icons.REMOVE)
            e.control.content.name = ft.Icons.REMOVE
            e.control.bgcolor = ft.Colors.RED
            self.snack_bar.content.value = f"{e.control.data[1]} added to quotation"
        else:
            remove_quote_row = [e.control.data[0], e.control.data[1], e.control.data[5], e.control.data[4]]
            if remove_quote_row in self.quotation_list:
                self.quotation_list.remove(remove_quote_row)
                self.page.client_storage.set("quotation", self.quotation_list)

            self.page.client_storage.set("add_remove_quotation_icon", ft.Icons.ADD)
            e.control.content.name = ft.Icons.ADD
            e.control.bgcolor = ft.Colors.GREEN
            self.snack_bar.content.value = f"{e.control.data[1]} removed from quotation"

        self.page.open(self.snack_bar)

        self.page.update()

    # clear quotation
    def clear_quotation(self, e):
        self.quotation_list.clear()
        self.page.client_storage.remove("quotation")

        # to set quotation btn to add icon
        for i in range(1, self.sh.max_row + 1):
            self.data_row.controls[i - 2].controls[0].content.content.name = ft.Icons.ADD
            self.data_row.controls[i - 2].controls[0].content.bgcolor = ft.Colors.GREEN

        self.snack_bar.content.value = "Quotation cleared"
        self.page.open(self.snack_bar)

        self.page.update()
