import time
import flet as ft
import openpyxl


class HomePage(ft.Column):
    # load excel workbook and select active sheet
    wb = openpyxl.load_workbook("products_db.xlsx")
    sh = wb.active

    # quotation list
    quotation_list = []

    # search box
    search_box = ft.CupertinoTextField(
        placeholder_text="Search",
        placeholder_style=ft.TextStyle(color=ft.Colors.GREY),
        color=ft.Colors.BLACK,
        bgcolor=ft.Colors.WHITE,
        border_radius=0,
        padding=ft.padding.only(
            left=20, top=15, bottom=15
        ),
        border=ft.border.only(
            bottom=ft.BorderSide(width=1, color=ft.Colors.GREY_300)
        ),
        suffix=ft.Container(
            ft.Icon(
                name=None,
                color=ft.Colors.RED
            ),
            bgcolor=ft.Colors.WHITE,
            padding=ft.padding.only(right=20),
            # on_click=this is assigned in __init__ function,
        ),
        # on_change=this is assigned in __init__ function
    )

    # product contents
    products_content = ft.Column(
        spacing=1,
        height=670,
        scroll=ft.ScrollMode.HIDDEN
    )

    percent_button_for_Sale_rate = ft.Container(
        content=ft.Icon(
            name=ft.Icons.PERCENT,
            color=ft.Colors.WHITE,
            size=18
        ),
        padding=5,
        border_radius=10,
        bgcolor=ft.Colors.GREEN,
        ink=True,
        # defined in __init function
        # on_click=self.set_sale_rate_percentage
    )

    # product heading
    products = ft.Row(
        spacing=1,
        scroll=ft.ScrollMode.HIDDEN,
        controls=[
            ft.Column(
                spacing=0,
                controls=[
                    ft.Row(
                        spacing=1,
                        controls=[
                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=80,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="Quote",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=80,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="Sr No",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=300,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="Goods Description",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=150,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="Part Number",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=150,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="Purchase Rate",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                ),
                                visible=False
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=150,
                                height=40,
                                padding=ft.padding.only(left=10, right=10),
                                alignment=ft.alignment.center,
                                # content=ft.Text(
                                #     value="Sale Rate",
                                #     color=ft.Colors.BLACK,
                                #     weight=ft.FontWeight.W_500
                                # ),
                                content=ft.Row(
                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                                    spacing=0,
                                    controls=[
                                        ft.Text(
                                            value="Sale Rate",
                                            color=ft.Colors.BLACK,
                                            weight=ft.FontWeight.W_500
                                        ),

                                        percent_button_for_Sale_rate
                                    ]
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=80,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="Stock",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=150,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="HSN Code",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=80,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="GST",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            ),

                            ft.Container(
                                bgcolor=ft.Colors.GREY_400,
                                width=150,
                                height=40,
                                alignment=ft.alignment.center,
                                content=ft.Text(
                                    value="Total",
                                    color=ft.Colors.BLACK,
                                    weight=ft.FontWeight.W_500
                                )
                            )
                        ]
                    ),

                    products_content
                ]

            )
        ]
    )

    def __init__(self, page, visibility_purchase_btn):
        super(HomePage, self).__init__()

        self.page = page

        # to access visibility btn from main.py
        self.visibility_purchase_btn = visibility_purchase_btn

        # assign on_click and on_change event to search box
        # assigned here because outside __init__ function we can not able to use self
        self.search_box.suffix.on_click = self.clear_search_field
        self.search_box.on_change = self.search_field_validation

        # assign on_click event to percent button which is used to calculate sale rate
        self.percent_button_for_Sale_rate.on_click = self.set_sale_rate_percentage

        # alert dialog box
        self.dialog_box = ft.AlertDialog(
            # dialog box background color
            surface_tint_color=ft.Colors.WHITE,
            scrollable=True,
            modal=True,
        )

        # snack bar
        self.snack_bar = ft.SnackBar(
            content=ft.Text()
        )

        self.spacing = 0
        self.controls = [
            # search box
            self.search_box,

            # products heading
            self.products
        ]

        # call the function to append products content
        self.add_product_content()

    # function to add products content
    def add_product_content(self):
        # first clear all the product contents
        self.products_content.controls.clear()

        for i in range(2, self.sh.max_row + 1):
            try:
                sale_percent = (self.page.client_storage.get("sale_percent") + 100) / 100
            except Exception as e:
                sale_percent = 1.1

            sale_rate = round(float(self.sh.cell(row=i, column=4).value) * sale_percent)
            total = round(float(sale_rate) * ((float(self.sh.cell(row=i, column=7).value) + float(100)) / float(100)))

            self.products_content.controls.append(
                # append products content
                ft.Row(
                    spacing=1,
                    controls=[
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            width=80,
                            alignment=ft.alignment.center,
                            content=ft.Icon(
                                name=ft.Icons.ADD,
                                color=ft.Colors.GREEN,
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Quote",  # column name
                                "1",  # column number in Excel file
                            ],
                            on_click=self.prepare_quotation
                        ),

                        # Sr No column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            width=80,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=1).value),
                                color=ft.Colors.BLACK,
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Sr No",  # column name
                                "1",  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # Goods Description column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=ft.padding.only(left=10, right=10),
                            width=300,
                            alignment=ft.alignment.center_left,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=2).value),
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Goods Description",  # column name
                                "2"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # Part Number column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=10,
                            width=150,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=3).value),
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Part Number",  # column name
                                "3"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # Purchase Rate column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=10,
                            width=150,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=f"{str(self.sh.cell(row=i, column=4).value)} \u20B9",
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Purchase Rate",  # column name
                                "4"  # column number in Excel file
                            ],
                            visible=False,
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # Sale Rate column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=10,
                            width=150,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=f"{str(sale_rate)} \u20B9",
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Sale Rate",  # column name
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # Stock column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=10,
                            width=80,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=5).value),
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Stock",  # column name
                                "5"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # HSN Code column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=10,
                            width=150,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=str(self.sh.cell(row=i, column=6).value),
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "HSN Code",  # column name
                                "6"  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # GST column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=10,
                            width=80,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=f"{str(self.sh.cell(row=i, column=7).value)}%",
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "GST",  # column name
                                "7",  # column number in Excel file
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        ),

                        # Total column
                        ft.Container(
                            bgcolor=ft.Colors.WHITE,
                            padding=10,
                            width=150,
                            alignment=ft.alignment.center,
                            content=ft.Text(
                                value=f"{str(total)} \u20B9",
                                color=ft.Colors.BLACK
                            ),
                            data=[
                                str(self.sh.cell(row=i, column=1).value),  # Sr No value
                                str(self.sh.cell(row=i, column=2).value),  # Goods Description value
                                str(self.sh.cell(row=i, column=3).value),  # Part number value
                                f"{self.sh.cell(row=i, column=4).value} \u20B9",  # Purchase Rate value
                                f"{sale_rate} \u20B9",  # Sale rate value
                                str(self.sh.cell(row=i, column=5).value),  # Stock value
                                str(self.sh.cell(row=i, column=6).value),  # HSN Code value
                                f"{self.sh.cell(row=i, column=7).value}%",  # GST value
                                f"{round(sale_rate * 0.18 + sale_rate)} \u20B9",  # Total value
                                "Total",  # column name
                            ],
                            on_click=self.view_record,
                            on_long_press=self.edit_record
                        )
                    ]
                )
            )

            # divider
            self.products_content.controls.append(
                ft.Container(
                    width=1228,
                    height=1,
                    bgcolor=ft.Colors.GREY_200
                )
            )

        self.already_quoted()

    # search validations
    def search_field_validation(self, e):
        # if search box is having data than close btn appears in the search box
        if e.control.value != "":
            e.control.suffix.content.name = ft.Icons.CLOSE_ROUNDED
        else:
            e.control.suffix.content.name = None

        # show only matched values with search box
        for i in self.products_content.controls:
            if self.products_content.controls.index(i) % 2 == 0:
                if str(self.search_box.value).lower() in str(i.controls[2].content.value).lower():
                    # visible products content
                    i.visible = True
                    # visible divider
                    self.products_content.controls[int(self.products_content.controls.index(i)) + 1].visible = True
                else:
                    # hide products content
                    i.visible = False
                    # hide products content
                    self.products_content.controls[int(self.products_content.controls.index(i)) + 1].visible = False

        self.page.update()

    # clear search field
    def clear_search_field(self, e):
        # when clear btn pressed search box clear the search box content and hide the close btn
        e.control.content.name = None
        self.search_box.value = ""

        self.add_product_content()

        self.page.update()

    # view record on single click of cell
    def view_record(self, e):
        self.dialog_box.title = ft.Text(
            value=f"Record No {e.control.data[0]}"
        )
        self.dialog_box.content = ft.Column(
            controls=[
                # Goods Description heading and value
                ft.Text(
                    value="Goods Description",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[1],
                    color=ft.Colors.GREY
                ),

                # Part Number heading and value
                ft.Text(
                    value="Part Number",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[2],
                    color=ft.Colors.GREY
                ),

                # Purchase Rate heading and value
                ft.Text(
                    value="Purchase Rate",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600,
                    visible=False
                ),
                ft.Text(
                    value=e.control.data[3],
                    color=ft.Colors.GREY,
                    visible=False
                ),

                # Sale Rate heading and value
                ft.Text(
                    value="Sale Rate",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[4],
                    color=ft.Colors.GREY
                ),

                # Stock heading and value
                ft.Text(
                    value="Stock",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[5],
                    color=ft.Colors.GREY
                ),

                # HSN Code heading and value
                ft.Text(
                    value="HSN Code",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[6],
                    color=ft.Colors.GREY
                ),

                # GST heading and value
                ft.Text(
                    value="GST",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[7],
                    color=ft.Colors.GREY
                ),

                # Total heading and value
                ft.Text(
                    value="Total",
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.W_600
                ),
                ft.Text(
                    value=e.control.data[8],
                    color=ft.Colors.GREY
                ),
            ]
        )

        # if purchase rate visibility is hidden than here also it is hidden else visible
        if self.visibility_purchase_btn.content.controls[0].name == ft.Icons.VISIBILITY_OFF:
            #  purchase rate heading visibility is hidden in view_record
            self.dialog_box.content.controls[4].visible = False
            #  purchase rate content visibility is hidden in view_record
            self.dialog_box.content.controls[5].visible = False
        else:
            #  purchase rate heading visibility is visible in view_record
            self.dialog_box.content.controls[4].visible = True
            #  purchase rate content visibility is visible in view_record
            self.dialog_box.content.controls[5].visible = True

        self.dialog_box.actions.clear()
        self.dialog_box.actions.append(
            ft.TextButton(
                text="Close",
                on_click=lambda e: self.page.close(self.dialog_box)
            )
        )
        self.page.open(self.dialog_box)

    def edit_record(self, e):
        # Sr No, Sale Rate and Total columns are not editable
        if e.control.data[9] == "Sr No" or e.control.data[9] == "Sale Rate" or e.control.data[9] == "Total":
            self.dialog_box.title = ft.Text(
                value=e.control.data[9]
            )

            self.dialog_box.content = ft.Text(
                value=f"You can't update {e.control.data[9]}"
            )

            self.dialog_box.actions.clear()
            self.dialog_box.actions.append(
                ft.TextButton(
                    text="Close",
                    on_click=lambda e: self.page.close(self.dialog_box)
                )
            )

            self.page.open(self.dialog_box)
        else:
            self.dialog_box.title = ft.Text(
                value=f"Update Record No {e.control.data[0]}"
            )

            self.dialog_box.title = ft.Row(
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                controls=[
                    ft.Text(
                        value=f"Update Record No {e.control.data[0]}"
                    ),

                    ft.IconButton(
                        icon=ft.Icons.DELETE,
                        icon_color=ft.Colors.RED,
                        data=e,
                        on_click=self.delete_record
                    )
                ]
            )

            # this is the value for purchase column to avoid rupee symbol in textfield of editing
            if e.control.data[9] == "Purchase Rate":
                self.dialog_box.content = ft.CupertinoTextField(
                    value=str(e.control.content.value)[0:str(e.control.content.value).find("\u20B9") - 1]
                )
            # this is the value for GST column to avoid % symbol in textfield of editing
            elif e.control.data[9] == "GST":
                self.dialog_box.content = ft.CupertinoTextField(
                    value=str(e.control.content.value).removesuffix("%")
                )
            # this is for all other columns
            else:
                self.dialog_box.content = ft.Column(
                    controls=[
                        ft.Text(
                            value=f"{e.control.data[9]}",
                            size=18,
                            weight=ft.FontWeight.W_500,
                            color=ft.Colors.GREY
                        ),

                        ft.Container(
                            expand=True,
                            content=ft.Text(
                                value=e.control.data[1],
                                weight=ft.FontWeight.W_500,
                                color=ft.Colors.GREY_500
                            )
                        ),

                        ft.CupertinoTextField(
                            value=str(e.control.content.value),
                            padding=ft.padding.only(left=10, right=10, top=15, bottom=15)
                        )
                    ]
                )

            self.dialog_box.actions.clear()
            self.dialog_box.actions.append(
                ft.TextButton(
                    text="Update",
                    data=e,
                    on_click=self.update_record
                )
            )
            self.dialog_box.actions.append(
                ft.TextButton(
                    text="Close",
                    on_click=lambda e: self.page.close(self.dialog_box)
                )
            )

            self.page.open(self.dialog_box)

    def delete_record(self, e):
        for i in range(2, self.sh.max_row + 1):
            if str(self.sh.cell(row=i, column=1).value) == str(e.control.data.control.data[0]):
                # deleting row
                self.sh.delete_rows(self.sh.cell(row=i, column=1).row)

                # saving workbook
                self.wb.save("products_db.xlsx")

                # to show deleted updates
                self.add_product_content()

                # closing dialog_box
                self.page.close(self.dialog_box)

                # notifying through snack bar
                self.snack_bar.content.value = f"Record Number {e.control.data.control.data[0]} deleted"
                self.page.open(self.snack_bar)

        self.page.update()

    def update_record(self, e):
        # update the cell
        if str(self.dialog_box.content.controls[2].value).isdigit():
            # if string contains only digits
            self.sh.cell(row=int(e.control.data.control.data[0]) + 1,
                         column=int(e.control.data.control.data[10])).value = int(self.dialog_box.content.controls[2].value)
        else:
            # if string contains alphanumeric
            self.sh.cell(row=int(e.control.data.control.data[0]) + 1,
                         column=int(e.control.data.control.data[10])).value = str(self.dialog_box.content.controls[2].value)

        # save Excel file
        self.wb.save("products_db.xlsx")

        # add done icon to show that record is updated
        self.dialog_box.actions.insert(
            0,
            ft.Icon(
                name=ft.Icons.DONE,
                color=ft.Colors.GREEN
            )
        )

        # to show updates call add_product_content function
        self.add_product_content()

        self.page.update()

        # wait for 1 second
        time.sleep(1)

        self.page.close(self.dialog_box)
        self.dialog_box.actions.clear()

    def prepare_quotation(self, e):
        quote_row = [e.control.data[1], 1, e.control.data[4]]
        if e.control.content.name == ft.Icons.ADD:
            e.control.content.name = ft.Icons.REMOVE
            e.control.content.color = ft.Colors.RED

            if quote_row not in self.quotation_list:
                self.quotation_list.append(quote_row)
        else:
            e.control.content.name = ft.Icons.ADD
            e.control.content.color = ft.Colors.GREEN

            if quote_row in self.quotation_list:
                self.quotation_list.remove(quote_row)

        self.page.update()

    # if already added in quotation list than add icon becomes remove icon or quote column
    def already_quoted(self):
        for i in range(0, self.sh.max_row * 2 - 2):
            if i % 2 == 0:
                quote_row = [self.products_content.controls[i].controls[0].data[1], 1,
                             self.products_content.controls[i].controls[0].data[4]]

                if quote_row in self.quotation_list:
                    self.products_content.controls[i].controls[0].content.name = ft.Icons.REMOVE
                    self.products_content.controls[i].controls[0].content.color = ft.Colors.RED

        self.page.update()

    def set_sale_rate_percentage(self, e):
        self.dialog_box.title = ft.Text("Percentage for Sale Rate")
        self.dialog_box.content = ft.TextField(
            hint_text="Percent",
            keyboard_type=ft.KeyboardType.NUMBER
        )
        self.dialog_box.actions.clear()
        self.dialog_box.actions = [
            ft.TextButton(
                text="Set",
                on_click=self.set_sale_percent
            ),

            ft.TextButton(
                text="Close",
                on_click=lambda e: self.page.close(self.dialog_box)
            )
        ]

        self.page.open(self.dialog_box)

    def set_sale_percent(self, e):
        self.page.client_storage.set("sale_percent", float(self.dialog_box.content.value))
        self.dialog_box.actions = [
            ft.Icon(
                name=ft.Icons.DONE,
                color=ft.Colors.GREEN
            ),

            ft.TextButton(
                text="Set",
                on_click=self.set_sale_percent
            ),

            ft.TextButton(
                text="Close",
                on_click=lambda e: self.page.close(self.dialog_box)
            )
        ]

        self.add_product_content()

        time.sleep(1)

        self.page.close(self.dialog_box)

        self.page.update()
